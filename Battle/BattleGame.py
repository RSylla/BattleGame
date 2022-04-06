import os
from openpyxl import load_workbook

import main

army_num = int(input("How many armies you want to create? "))
size = int(input("How many soldiers armies contain? "))

for x in range(army_num):
    army_name = input(f"Please insert the name for army No {x+1}: ")
    army = main.Army(size,army_name)
    army.army_generator()

attr_value = 0
attr_key = 6
armies_list = os.listdir("Armies/")
directory = "Armies/"

def army_vs_army(size,index):
    army_1_name = armies_list[index][:-5]
    if index != len(armies_list)-1:
        army_2_name = armies_list[index+1][:-5]
    else:
        army_2_name = armies_list[0][:-5]
    army_1_pts = 0
    army_2_pts = 0
    count = size + 1
    for x in range(1,count):
        fight = one_vs_one(count,army_1_name,army_2_name)
        # print("------------------------------------------------")
        count -= 1
        if fight == "Army1":
            army_1_pts += 1
        elif fight == "Army2":
            army_2_pts += 1
    if army_1_pts > army_2_pts:
        return f"{army_1_name} won the battle against {army_2_name}. \nWith points: {army_1_pts}:{army_2_pts} \n\n"
    if army_1_pts < army_2_pts:
        return f"{army_2_name} won the battle against {army_1_name}. \nWith points: {army_2_pts}:{army_1_pts} \n\n"

def one_vs_one(size,name1, name2):
    key = 4
    opponent_1_pts = 0
    opponent_2_pts = 0
    opponent_1_name = wb_sheet.cell(size,2).value
    opponent_2_name = wb2_sheet.cell(size,2).value
    opponent_1_type = wb_sheet.cell(size,3).value
    opponent_2_type = wb2_sheet.cell(size,3).value
    # print(f"{name1}'s {opponent_1_name}({opponent_1_type}) and {name2}'s {opponent_2_name}({opponent_2_type}) are fighting!")

    for x in range(4,9):
        opp1 = wb_sheet.cell(size,key).value
        opp2 = wb2_sheet.cell(size,key).value
        attr = wb_sheet.cell(1,key).value
        a = ""
        if opp1 > opp2:
            opponent_1_pts += 1
            a = f"{opponent_1_name.split()[0]} gets the point!"
        elif opp2 > opp1:
            opponent_2_pts += 1
            a = f"{opponent_2_name.split()[0]} gets the point!"
        else:
            a = "It was a tie!"
        key += 1
        # print(f"{attr}: {opp1} vs {opp2} --- {a}")

    if opponent_1_pts > opponent_2_pts:
        # print(f"{opponent_1_pts}:{opponent_2_pts} {opponent_1_name.split()[0]} won this battle!")
        return "Army1"
    elif opponent_2_pts > opponent_1_pts:
        # print(f"{opponent_2_pts}:{opponent_1_pts} {opponent_2_name.split()[0]} won this battle!")
        return "Army2"

for index in range(len(armies_list)):
    if index != len(armies_list)-1:
        wb = load_workbook(f"{directory}" + f"{armies_list[index]}")
        wb_sheet = wb.active
        wb2 = load_workbook(f"{directory}" + f"{armies_list[index+1]}")
        wb2_sheet = wb2.active
    if index == len(armies_list)-1:
        wb = load_workbook(f"{directory}" + f"{armies_list[-1]}")
        wb_sheet = wb.active
        wb2 = load_workbook(f"{directory}" + f"{armies_list[0]}")
        wb2_sheet = wb2.active

    print(army_vs_army(size,index))

